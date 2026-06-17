"""
微信公众号文章导出工具

支持将文章数据导出为多种格式：
- JSON: 完整结构化数据
- Excel: 表格形式，适合筛选分析
- Markdown: 从 HTML 转换，适合阅读和编辑
- TXT: 纯文本，适合全文搜索
- HTML: 保存原始 HTML 文件（需先下载）
"""

import json
import os
import re
from datetime import datetime
from typing import Optional

# 延迟导入可选依赖
_openpyxl = None
_markdownify = None
_bs4 = None


def _get_openpyxl():
    global _openpyxl
    if _openpyxl is None:
        import openpyxl
        _openpyxl = openpyxl
    return _openpyxl


def _get_markdownify():
    global _markdownify
    if _markdownify is None:
        import markdownify
        _markdownify = markdownify
    return _markdownify


def _get_bs4():
    global _bs4
    if _bs4 is None:
        from bs4 import BeautifulSoup
        _bs4 = BeautifulSoup
    return _bs4


def sanitize_filename(name: str, max_length: int = 80) -> str:
    """清理文件名"""
    name = re.sub(r'[\\/:*?"<>|\n\r\t]', '_', name)
    name = name.strip('. ')
    if len(name) > max_length:
        name = name[:max_length]
    return name or "untitled"


class ArticleExporter:
    """文章导出器"""

    def __init__(self, articles: list, output_dir: str, account_name: str = ""):
        """
        Args:
            articles: 文章列表（从 WeChatMPClient.fetch_all_articles 获取）
            output_dir: 输出根目录
            account_name: 公众号名称（用于文件命名）
        """
        self.articles = articles
        self.output_dir = output_dir
        self.account_name = account_name
        os.makedirs(output_dir, exist_ok=True)

    def _format_time(self, timestamp: int) -> str:
        """将 Unix 时间戳格式化为可读日期"""
        if not timestamp:
            return ""
        return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")

    def _format_date(self, timestamp: int) -> str:
        """将 Unix 时间戳格式化为日期（仅日期部分）"""
        if not timestamp:
            return "unknown"
        return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d")

    # ============================================================
    # JSON 导出
    # ============================================================

    def export_json(self, filename: str = "articles.json", indent: int = 2) -> str:
        """
        导出为 JSON 文件

        Args:
            filename: 输出文件名
            indent: JSON 缩进

        Returns:
            输出文件路径
        """
        output_path = os.path.join(self.output_dir, filename)

        export_data = {
            "account_name": self.account_name,
            "total_count": len(self.articles),
            "export_time": datetime.now().isoformat(),
            "articles": []
        }

        for article in self.articles:
            export_data["articles"].append({
                "title": article.get("title", ""),
                "link": article.get("link", ""),
                "digest": article.get("digest", ""),
                "cover": article.get("cover", ""),
                "create_time": article.get("create_time", 0),
                "create_time_str": self._format_time(article.get("create_time", 0)),
                "update_time": article.get("update_time", 0),
                "aid": article.get("aid", ""),
            })

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(export_data, f, ensure_ascii=False, indent=indent)

        print(f"已导出 JSON: {output_path} ({len(self.articles)} 篇)")
        return output_path

    # ============================================================
    # Excel 导出
    # ============================================================

    def export_excel(self, filename: str = "articles.xlsx") -> str:
        """
        导出为 Excel 文件

        Args:
            filename: 输出文件名

        Returns:
            输出文件路径
        """
        openpyxl = _get_openpyxl()
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        output_path = os.path.join(self.output_dir, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "文章列表"

        # 表头
        headers = ["序号", "标题", "发布日期", "摘要", "链接"]
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # 数据行
        for i, article in enumerate(self.articles):
            row = i + 2
            ws.cell(row=row, column=1, value=i + 1).border = thin_border
            ws.cell(row=row, column=2, value=article.get("title", "")).border = thin_border
            ws.cell(
                row=row, column=3,
                value=self._format_time(article.get("create_time", 0))
            ).border = thin_border
            ws.cell(row=row, column=4, value=article.get("digest", "")).border = thin_border

            link_cell = ws.cell(row=row, column=5, value=article.get("link", ""))
            link_cell.border = thin_border
            link = article.get("link", "")
            if link:
                link_cell.hyperlink = link
                link_cell.font = Font(color="0563C1", underline="single")

        # 列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 30

        # 冻结首行
        ws.freeze_panes = "A2"

        wb.save(output_path)
        print(f"已导出 Excel: {output_path} ({len(self.articles)} 篇)")
        return output_path

    # ============================================================
    # Markdown 导出（从已下载的 HTML 转换）
    # ============================================================

    def export_markdown(self, html_dir: Optional[str] = None) -> str:
        """
        将已下载的 HTML 文章转换为 Markdown

        Args:
            html_dir: HTML 文件所在目录，默认 {output_dir}/html

        Returns:
            Markdown 输出目录路径
        """
        markdownify = _get_markdownify()
        BeautifulSoup = _get_bs4()

        if html_dir is None:
            html_dir = os.path.join(self.output_dir, "html")

        md_dir = os.path.join(self.output_dir, "markdown")
        os.makedirs(md_dir, exist_ok=True)

        if not os.path.exists(html_dir):
            print(f"HTML 目录不存在: {html_dir}，请先下载文章 HTML")
            return md_dir

        converted = 0
        for filename in sorted(os.listdir(html_dir)):
            if not filename.endswith(".html"):
                continue

            html_path = os.path.join(html_dir, filename)
            md_filename = filename.replace(".html", ".md")
            md_path = os.path.join(md_dir, md_filename)

            try:
                with open(html_path, "r", encoding="utf-8") as f:
                    html_content = f.read()

                # 提取正文区域
                soup = BeautifulSoup(html_content, "html.parser")
                content_div = soup.find("div", {"id": "js_content"}) or soup.find("div", class_="rich_media_content")

                if content_div:
                    # 提取标题
                    title_el = soup.find("h1", class_="rich_media_title") or soup.find("h2", class_="rich_media_title")
                    title = title_el.get_text(strip=True) if title_el else filename.replace(".html", "")

                    md_content = f"# {title}\n\n"
                    md_content += markdownify.markdownify(
                        str(content_div),
                        heading_style="ATX",
                        strip=["script", "style", "iframe"],
                    )
                else:
                    md_content = markdownify.markdownify(
                        html_content,
                        heading_style="ATX",
                        strip=["script", "style", "iframe"],
                    )

                with open(md_path, "w", encoding="utf-8") as f:
                    f.write(md_content)

                converted += 1

            except Exception as e:
                print(f"  转换失败: {filename} - {e}")

        print(f"已导出 Markdown: {md_dir} ({converted} 篇)")
        return md_dir

    # ============================================================
    # TXT 导出（从已下载的 HTML 提取纯文本）
    # ============================================================

    def export_txt(self, html_dir: Optional[str] = None) -> str:
        """
        将已下载的 HTML 文章提取为纯文本

        Args:
            html_dir: HTML 文件所在目录

        Returns:
            TXT 输出目录路径
        """
        BeautifulSoup = _get_bs4()

        if html_dir is None:
            html_dir = os.path.join(self.output_dir, "html")

        txt_dir = os.path.join(self.output_dir, "txt")
        os.makedirs(txt_dir, exist_ok=True)

        if not os.path.exists(html_dir):
            print(f"HTML 目录不存在: {html_dir}，请先下载文章 HTML")
            return txt_dir

        converted = 0
        for filename in sorted(os.listdir(html_dir)):
            if not filename.endswith(".html"):
                continue

            html_path = os.path.join(html_dir, filename)
            txt_filename = filename.replace(".html", ".txt")
            txt_path = os.path.join(txt_dir, txt_filename)

            try:
                with open(html_path, "r", encoding="utf-8") as f:
                    html_content = f.read()

                soup = BeautifulSoup(html_content, "html.parser")

                # 移除 script 和 style
                for tag in soup(["script", "style"]):
                    tag.decompose()

                # 提取标题
                title_el = soup.find("h1", class_="rich_media_title") or soup.find("h2", class_="rich_media_title")
                title = title_el.get_text(strip=True) if title_el else ""

                # 提取正文
                content_div = soup.find("div", {"id": "js_content"}) or soup.find("div", class_="rich_media_content")
                if content_div:
                    text = content_div.get_text(separator="\n", strip=True)
                else:
                    text = soup.get_text(separator="\n", strip=True)

                full_text = f"{title}\n\n{text}" if title else text

                with open(txt_path, "w", encoding="utf-8") as f:
                    f.write(full_text)

                converted += 1

            except Exception as e:
                print(f"  转换失败: {filename} - {e}")

        print(f"已导出 TXT: {txt_dir} ({converted} 篇)")
        return txt_dir

    # ============================================================
    # 便捷方法：全格式导出
    # ============================================================

    def export_all(self, html_dir: Optional[str] = None) -> dict:
        """
        导出所有支持的格式

        Returns:
            各格式的输出路径字典
        """
        results = {}

        # JSON 和 Excel 只需要文章元数据
        results["json"] = self.export_json()
        results["excel"] = self.export_excel()

        # Markdown 和 TXT 需要已下载的 HTML
        if html_dir or os.path.exists(os.path.join(self.output_dir, "html")):
            results["markdown"] = self.export_markdown(html_dir)
            results["txt"] = self.export_txt(html_dir)
        else:
            print("\n提示: Markdown 和 TXT 导出需要先下载文章 HTML")

        return results
